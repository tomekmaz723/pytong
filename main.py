import pandas as pd
import openpyxl
from Bus import Bus
from collections import Counter


# df = pd.read_excel(r'D:\OneDrive\Pulpit\projekt.xlsx')


# data = list(df['Imie'])
# mylist = df['Nazwisko'].tolist()
# print(data)
# pdw
#

def load(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    cell_range = 'A2:D76'
    data = [[cell.value for cell in row] for row in ws[cell_range]]
    print(data)


# sortowanie listy po zabawometrze

def zabawometr():
    data.sort(key=lambda x: x[3], reverse=True)


# sortowanie listy po występowaniu kierunku

def kierunki():
    df = pd.DataFrame(data, columns=['First Name', 'Surname', 'Kierunek', 'Zabawa'])
    df['Frequency'] = df.groupby('Kierunek')['Kierunek'].transform('count')
    df.sort_values('Frequency', inplace=True, ascending=False)
    lol = df.values.tolist()
    print(lol)
    for i in range(len(lol)):
        lol[i].pop(4)
    return lol


# ładowanie danych do listy

wb = openpyxl.load_workbook('projekt.xlsx')
ws = wb.active
cell_range = 'A2:D76'  # zakres
data = [[cell.value for cell in row] for row in ws[cell_range]]
data1 = [[cell.value for cell in row] for row in ws[cell_range]]

# df['count'] = df.groupby('Kierunek')['Kierunek'].transform(pd.Series.value_counts)
# df.sort_values('count', ascending=False)
# kierunki = df.values.tolist()
# https://stackoverflow.com/questions/30787391/sorting-entire-csv-by-frequency-of-occurence-in-one-column

print(kierunki())

print(zabawometr())

# autobusy

# x = int(input("miejsca w busie: "))
x = 30
y = x - 1
z = x - 2
q = x - 0
b1 = Bus("b1", x, y)
b2 = Bus("b2", x, z)
b3 = Bus("b3", x, q)
buses = [b1, b2, b3]
buses.sort(key=lambda x: x.empty_seat, reverse=True)
print(buses)

print(data)

list_imie = []
list_nazwisko = []

# dodwanie imion i nazwisk # # # # # # # # # # # #

for i in range(20):
    list_imie.append(data[i][0])
    list_nazwisko.append(data[i][1])

dymy = [list(a) for a in zip(list_imie, list_nazwisko)]  # merge lists of first name and surnames
print(dymy)

# # # # # # # # # # # # # # # # # # # # # # # #



for i in buses:
    empty = i.empty_seat
    kierunek = data[0][2]
    tcount = sum(x[2] == kierunek for x in data)
